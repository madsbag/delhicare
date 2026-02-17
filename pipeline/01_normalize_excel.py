#!/usr/bin/env python3
"""
Step 1: Normalize the classified Excel into a clean JSON for the website pipeline.

Reads rehab_care_classified.xlsx, filters to 182 usable businesses (removing
UNKNOWN and UNCERTAIN), corrects INCORRECT categories using Inferred Category,
normalizes cities, generates URL slugs, and outputs normalized_businesses.json.
"""

import json
import re
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

# ─── Configuration ────────────────────────────────────────────────────────────
CLASSIFIED_XLSX = Path("/Users/madhur/Downloads/rehab_care_classified.xlsx")
OUTPUT_DIR = Path(__file__).parent / "data"
OUTPUT_FILE = OUTPUT_DIR / "normalized_businesses.json"

# Categories we keep for the directory
DIRECTORY_CATEGORIES = [
    "Nursing Homes",
    "Elder Care",
    "Post-Hospital Care",
    "Home Health Care",
]

# Rename mapping: original category -> display category
CATEGORY_RENAME = {
    "Rehabilitation": "Post-Hospital Care",
}

# Category match values to keep (CORRECT + INCORRECT with known corrections)
KEEP_MATCH_VALUES = ["CORRECT", "INCORRECT"]

# City normalization map (Outscraper data has inconsistent city names)
CITY_NORMALIZATION = {
    # Delhi variants
    "new delhi": "Delhi",
    "delhi": "Delhi",
    "south delhi": "Delhi",
    "north delhi": "Delhi",
    "east delhi": "Delhi",
    "west delhi": "Delhi",
    "central delhi": "Delhi",
    "south west delhi": "Delhi",
    "north west delhi": "Delhi",
    "north east delhi": "Delhi",
    "south east delhi": "Delhi",
    "shahdara": "Delhi",
    "dwarka": "Delhi",
    "rohini": "Delhi",
    "pitampura": "Delhi",
    "janakpuri": "Delhi",
    "lajpat nagar": "Delhi",
    "karol bagh": "Delhi",
    "saket": "Delhi",
    "vasant kunj": "Delhi",
    "defence colony": "Delhi",
    "greater kailash": "Delhi",
    # Gurgaon/Gurugram variants
    "gurugram": "Gurgaon",
    "gurgaon": "Gurgaon",
    # Noida variants
    "noida": "Noida",
    "greater noida": "Greater Noida",
    # Ghaziabad variants
    "ghaziabad": "Ghaziabad",
    # Faridabad variants
    "faridabad": "Faridabad",
    # Meerut
    "meerut": "Meerut",
    # Lucknow (some may spill from UP data)
    "lucknow": "Lucknow",
    # Sonipat / Panipat
    "sonipat": "Sonipat",
    "panipat": "Panipat",
    # Bahadurgarh
    "bahadurgarh": "Bahadurgarh",
    # Hapur
    "hapur": "Hapur",
    # Other NCR areas
    "border": "Delhi",
    "trun": "Delhi",
    "dadri toye": "Greater Noida",
    "aurangpur": "Greater Noida",
    "bisrakh jalalpur": "Greater Noida",
    "sarhol": "Gurgaon",
    "kadarpur": "Gurgaon",
    "mahendwara": "Gurgaon",
    "afzalpur": "Ghaziabad",
    "asalat nagar": "Ghaziabad",
    "jawli": "Ghaziabad",
    "basi bahuddin nagar": "Noida",
    "pali village": "Faridabad",
    "sirohi": "Sirohi",
}

# Map inferred categories back to our directory categories
INFERRED_TO_DIRECTORY = {
    "Nursing Homes": "Nursing Homes",
    "Elder Care": "Elder Care",
    "Rehabilitation": "Post-Hospital Care",
    "Post-Hospital Care": "Post-Hospital Care",
    "Home Health Care": "Home Health Care",
    # Common LLM inferred names that may differ slightly
    "Nursing Home": "Nursing Homes",
    "Elderly Care": "Elder Care",
    "Rehabilitation Center": "Post-Hospital Care",
    "Home Health": "Home Health Care",
    "Home Healthcare": "Home Health Care",
    "Home Care": "Home Health Care",
    "Rehab": "Post-Hospital Care",
    "Rehab Center": "Post-Hospital Care",
    "Physical Rehabilitation": "Post-Hospital Care",
    # Other inferred categories that don't map to our 4
    "Other": None,
    "Hospitals & Medical Centers": None,
    "De-Addiction & Substance Abuse": None,
    "Mental Health & Psychiatric": None,
    "Physiotherapy & Physical Therapy": None,
    "Doctors & Specialists": None,
    "Wellness & Alternative Medicine": None,
    "Speech & Occupational Therapy": None,
}


def slugify(text: str) -> str:
    """Generate a URL-friendly slug from text."""
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[-\s]+", "-", text)
    text = text.strip("-")
    return text


def category_slug(category: str) -> str:
    """Generate a URL slug for a category."""
    return slugify(category)


def city_slug(city: str) -> str:
    """Generate a URL slug for a city."""
    return slugify(city)


def normalize_city(city_raw: str) -> str:
    """Normalize city name using the mapping.

    Handles compound city names like 'New Delhi, Delhi' or 'Gurugram, Sarhol'
    by checking each part and preferring the first match.
    """
    if not city_raw or not isinstance(city_raw, str):
        return "Delhi"  # Default to Delhi for missing city

    city_lower = city_raw.strip().lower()

    # Direct match first
    if city_lower in CITY_NORMALIZATION:
        return CITY_NORMALIZATION[city_lower]

    # Try each comma-separated part
    parts = [p.strip().lower() for p in city_lower.split(",")]
    for part in parts:
        if part in CITY_NORMALIZATION:
            return CITY_NORMALIZATION[part]

    # Try substring matching for compound names containing known cities
    for key, normalized in CITY_NORMALIZATION.items():
        if key in city_lower:
            return normalized

    # Default: title case the first part
    return parts[0].strip().title() if parts else "Delhi"


def parse_working_hours(hours_str: str) -> dict:
    """Parse pipe-delimited working hours into structured format.
    Example input: 'Monday: 9 AM–6 PM | Tuesday: 9 AM–6 PM | ...'
    """
    if not hours_str or not isinstance(hours_str, str):
        return {}
    hours = {}
    parts = hours_str.split("|")
    for part in parts:
        part = part.strip()
        if ":" in part:
            day, _, time_range = part.partition(":")
            hours[day.strip()] = time_range.strip()
    return hours


def extract_lat_lng(maps_link: str) -> tuple:
    """Extract lat/lng from Google Maps link.
    Example: https://www.google.com/maps/place/.../@28.6139,77.2090,...
    """
    if not maps_link or not isinstance(maps_link, str):
        return None, None
    match = re.search(r"@(-?\d+\.?\d*),(-?\d+\.?\d*)", maps_link)
    if match:
        return float(match.group(1)), float(match.group(2))
    return None, None


def clean_email(email: str) -> str:
    """Clean email addresses (strip artifacts like u003e prefix)."""
    if not email or not isinstance(email, str):
        return ""
    email = email.strip()
    # Remove common artifacts
    email = re.sub(r"^u003[ce]", "", email)
    email = re.sub(r"u003[ce]$", "", email)
    email = email.strip("<>")
    return email if "@" in email else ""


def normalize_phone(phone) -> str:
    """Normalize phone number to readable format."""
    if not phone:
        return ""
    phone = str(phone).strip()
    # Remove non-digit except +
    digits = re.sub(r"[^\d+]", "", phone)
    if not digits:
        return ""
    # Add +91 prefix if missing and looks like Indian number
    if digits.startswith("91") and len(digits) == 12:
        digits = "+" + digits
    elif digits.startswith("+91"):
        pass  # already correct
    elif len(digits) == 10:
        digits = "+91" + digits
    return digits


def read_classified_excel(path: Path) -> list:
    """Read the Classification Results sheet from the classified Excel."""
    wb = openpyxl.load_workbook(path)
    ws = wb["Classification Results"]

    headers = [cell.value for cell in ws[1]]
    businesses = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        businesses.append(record)

    print(f"Read {len(businesses)} businesses from Classification Results sheet")
    return businesses


def filter_and_correct(businesses: list) -> list:
    """Filter to CORRECT+INCORRECT, correct INCORRECT categories."""
    filtered = []
    stats = Counter()

    for biz in businesses:
        match = biz.get("Category Match", "")
        sheet_cat = biz.get("Sheet Category", "")
        inferred_cat = biz.get("Inferred Category", "")

        # Only keep CORRECT and INCORRECT
        if match not in KEEP_MATCH_VALUES:
            stats[f"removed_{match}"] += 1
            continue

        # Determine the final category
        if match == "CORRECT":
            final_category = CATEGORY_RENAME.get(sheet_cat, sheet_cat)
        else:
            # INCORRECT: use inferred category, mapped to our directory categories
            mapped = INFERRED_TO_DIRECTORY.get(inferred_cat)
            if mapped:
                final_category = mapped
                stats["corrected"] += 1
            else:
                # Inferred category doesn't map to our 4 directory categories
                # Keep original sheet category but flag it
                final_category = CATEGORY_RENAME.get(sheet_cat, sheet_cat)
                stats["kept_original_despite_incorrect"] += 1

        # Verify final category is one of our 4
        if final_category not in DIRECTORY_CATEGORIES:
            stats[f"removed_non_directory_{final_category}"] += 1
            continue

        biz["_final_category"] = final_category
        filtered.append(biz)
        stats[f"kept_{match}"] += 1

    print(f"\nFiltering results:")
    for k, v in sorted(stats.items()):
        print(f"  {k}: {v}")
    print(f"  Total kept: {len(filtered)}")

    return filtered


def normalize_business(biz: dict, slug_set: set) -> dict:
    """Normalize a single business record into the website data format."""
    name = biz.get("Business Name", "").strip()
    city = normalize_city(biz.get("City", ""))
    state = biz.get("State", "")
    category = biz.get("_final_category", "")

    # Generate slug with collision detection
    base_slug = slugify(f"{name}-{city}")
    slug = base_slug
    counter = 1
    while slug in slug_set:
        slug = f"{base_slug}-{counter}"
        counter += 1
    slug_set.add(slug)

    # Extract lat/lng
    lat, lng = extract_lat_lng(biz.get("Google Maps Link", ""))

    # Parse working hours
    hours = parse_working_hours(biz.get("Working Hours", ""))

    # Clean contact info
    email1 = clean_email(biz.get("Email", ""))
    email2 = clean_email(biz.get("Email 2", ""))
    phone = normalize_phone(biz.get("Phone", ""))
    phone2 = normalize_phone(biz.get("Phone 2", ""))
    phone3 = normalize_phone(biz.get("Phone 3", ""))

    # Website URL
    website = biz.get("Website", "")
    if website and isinstance(website, str):
        website = website.strip()
        if not website.startswith("http"):
            website = ""
    else:
        website = ""

    # Social media
    def clean_url(val):
        if val and isinstance(val, str) and val.strip().startswith("http"):
            return val.strip()
        return ""

    return {
        "slug": slug,
        "name": name,
        "category": category,
        "category_slug": category_slug(category),
        "google_category": biz.get("Google Category", ""),
        "city": city,
        "city_slug": city_slug(city),
        "state": state if isinstance(state, str) else "",
        "full_address": biz.get("Full Address", "") or "",
        "phone": phone,
        "phone2": phone2,
        "phone3": phone3,
        "email": email1,
        "email2": email2,
        "website": website,
        "rating": float(biz["Rating"]) if biz.get("Rating") else None,
        "reviews": int(biz["Reviews"]) if biz.get("Reviews") else 0,
        "verified": bool(biz.get("Verified")),
        "working_hours": hours,
        "lat": lat,
        "lng": lng,
        "google_maps_link": biz.get("Google Maps Link", "") or "",
        "facebook": clean_url(biz.get("Facebook", "")),
        "instagram": clean_url(biz.get("Instagram", "")),
        "linkedin": clean_url(biz.get("LinkedIn", "")),
        "twitter": clean_url(biz.get("Twitter", "")),
        "whatsapp": biz.get("WhatsApp", "") or "",
        "contact_name": biz.get("Contact Name", "") or "",
        "contact_title": biz.get("Contact Title", "") or "",
        # Classification data (for pipeline use)
        "original_sheet_category": biz.get("Sheet Category", ""),
        "category_match": biz.get("Category Match", ""),
        "inferred_category": biz.get("Inferred Category", ""),
        "classification_keywords": biz.get("Keywords", "") or "",
        "business_summary": biz.get("Business Summary", "") or "",
        # These will be filled by later pipeline steps
        "services": [],
        "specializations": [],
        "doctors": [],
        "facilities": [],
        "testimonials": [],
        "faqs": [],
        "about_description": "",
        "long_description": "",
        "meta_description": "",
        "seo_title": "",
    }


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Step 1: Read classified Excel
    businesses = read_classified_excel(CLASSIFIED_XLSX)

    # Step 2: Filter and correct categories
    filtered = filter_and_correct(businesses)

    # Step 3: Normalize each business
    slug_set = set()
    normalized = []
    for biz in filtered:
        normalized.append(normalize_business(biz, slug_set))

    # Step 4: Print summary stats
    print(f"\n{'='*60}")
    print(f"NORMALIZATION COMPLETE")
    print(f"{'='*60}")
    print(f"Total businesses: {len(normalized)}")

    cat_counts = Counter(b["category"] for b in normalized)
    print(f"\nBy category:")
    for cat, count in cat_counts.most_common():
        print(f"  {cat}: {count}")

    city_counts = Counter(b["city"] for b in normalized)
    print(f"\nBy city:")
    for city, count in city_counts.most_common():
        print(f"  {city}: {count}")

    with_website = sum(1 for b in normalized if b["website"])
    with_phone = sum(1 for b in normalized if b["phone"])
    with_email = sum(1 for b in normalized if b["email"])
    with_coords = sum(1 for b in normalized if b["lat"])
    print(f"\nData completeness:")
    print(f"  With website: {with_website}")
    print(f"  With phone: {with_phone}")
    print(f"  With email: {with_email}")
    print(f"  With coordinates: {with_coords}")

    # Step 5: Save
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(normalized, f, indent=2, ensure_ascii=False)

    print(f"\nOutput saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
